#!/usr/bin/env python3
"""
SCRIPT DE INGESTÃO MENSAL
Automatiza: validação → integração de OS → recálculo do RELATORIO

Uso:
    python atualizar_mes.py OS_FEVEREIRO.xlsm
    python atualizar_mes.py OS_MARCO.xlsx

O que faz:
    1. Lê o arquivo de OS do mês
    2. Valida colunas e formatos
    3. Padroniza ASSUNTO usando config (DE→PARA)
    4. Remove duplicados (OS já existentes na base)
    5. Integra novas OS na aba OS da planilha
    6. Recalcula a aba RELATORIO (replica XLOOKUPs)
    7. Gera relatório de integração
"""

import sys
import os
import pandas as pd
import openpyxl
from datetime import datetime


# Caminho do arquivo base (na raiz do projeto)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, 'NFxPRODUTO__1_.xlsx')

# Colunas esperadas na OS
COLUNAS_OS = [
    'ID _Ordem de Serviço', 'data_abertura_OS', 'data_fechamento_OS',
    'Descrição Assunto', 'ID_cliente', 'Razão', 'Almoxarifado',
    'id_produto', 'descricao_produto', 'id_patrimonio',
    'numero_patrimonial', 'numero_serie', 'status_comodato',
    'ASSUNTO PADRONIZADO'
]

# Mapeamento de colunas alternativas (caso arquivo venha com nomes diferentes)
COLUNAS_MAP = {
    'almox.descricao': 'Almoxarifado',
    'almox_descricao': 'Almoxarifado',
    'Almox.descricao': 'Almoxarifado',
}


def log(msg):
    print(f"  {msg}")


def validar_arquivo(filepath):
    """Valida se o arquivo existe e é legível."""
    if not os.path.exists(filepath):
        print(f"ERRO: Arquivo não encontrado: {filepath}")
        sys.exit(1)

    ext = os.path.splitext(filepath)[1].lower()
    if ext not in ('.xlsx', '.xlsm', '.xls', '.csv'):
        print(f"ERRO: Formato não suportado: {ext}")
        sys.exit(1)

    return ext


def ler_os_novo(filepath, ext):
    """Lê o arquivo de OS do mês."""
    if ext == '.csv':
        df = pd.read_csv(filepath)
    else:
        df = pd.read_excel(filepath)

    log(f"Arquivo lido: {len(df)} linhas, {len(df.columns)} colunas")
    log(f"Colunas: {list(df.columns)}")
    return df


# Padronização canônica de ASSUNTOS (mesma referência de dados.py)
PADRONIZACAO_ASSUNTO = {
    'INSTALACAO INTERNET': 'INSTALACAO',
    'Instalação Internet (descontinuado)': 'INSTALACAO',
    'Instalação Repetidor Wirelles': 'MESH',
    'Instalação Serviço Cortesia': 'INSTALACAO CORTESIA',
    'Instalação de Telefone': 'INSTALACAO TELEFONE',
    'MANUTENCAO DE REDE': 'MANUTENCAO',
    'MANUTENÇÃO TÉCNICA': 'MANUTENCAO',
    'MUDANÇA DE ENDEREÇO': 'MUDANCA ENDEREÇO',
    'SERVIÇOS TÉCNICOS DIVERSOS': 'MESH',
    'UPGRADE - EQUIPAMENTO': 'UPGRADE',
    '0.1.4 RETIRADA DE REPETIDOR WIRELESS': 'RETIRADA REPETIDOR',
    '0.1.5 RETIRADA ORDEM DE COLETA': 'RETIRADA COLETA',
    '0.1.6 RETIRADA PONTO DE INTERNET': 'RETIRADA DE PONTO',
    'RETIRADA ORDEM DE COLETA': 'RETIRADA COLETA',
    'RETIRADA PONTO DE INTERNET': 'RETIRADA DE PONTO',
}


def padronizar_colunas(df_novo, config):
    """Renomeia colunas alternativas e gera ASSUNTO PADRONIZADO."""
    # Renomear colunas conhecidas
    for col_orig, col_dest in COLUNAS_MAP.items():
        if col_orig in df_novo.columns and col_dest not in df_novo.columns:
            df_novo = df_novo.rename(columns={col_orig: col_dest})
            log(f"Coluna renomeada: '{col_orig}' → '{col_dest}'")

    # Gerar ASSUNTO PADRONIZADO se não existir
    if 'ASSUNTO PADRONIZADO' not in df_novo.columns:
        if 'Descrição Assunto' in df_novo.columns:
            assunto_map = dict(zip(config['DE'].dropna(), config['PARA'].dropna()))
            df_novo['ASSUNTO PADRONIZADO'] = (
                df_novo['Descrição Assunto']
                .map(assunto_map)
                .fillna('****')
            )
            mapeados = df_novo['ASSUNTO PADRONIZADO'].ne('****').sum()
            log(f"ASSUNTO PADRONIZADO gerado: {mapeados}/{len(df_novo)} mapeados")
        else:
            print("ERRO: Coluna 'Descrição Assunto' não encontrada. Impossível gerar ASSUNTO PADRONIZADO.")
            sys.exit(1)

    # Re-padronizar: aplica nomenclatura canônica sobre valores existentes
    df_novo['ASSUNTO PADRONIZADO'] = df_novo['ASSUNTO PADRONIZADO'].replace(PADRONIZACAO_ASSUNTO)
    log(f"Padronizacao canonica aplicada ao ASSUNTO PADRONIZADO")

    return df_novo


def validar_colunas_obrigatorias(df_novo):
    """Verifica se as colunas necessárias existem."""
    obrigatorias = ['ID _Ordem de Serviço', 'data_fechamento_OS', 'id_patrimonio']
    faltando = [c for c in obrigatorias if c not in df_novo.columns]
    if faltando:
        print(f"ERRO: Colunas obrigatórias faltando: {', '.join(faltando)}")
        print(f"Colunas disponíveis: {list(df_novo.columns)}")
        sys.exit(1)


def integrar_os(df_novo, df_base):
    """Remove duplicados e integra novas OS."""
    # Remover duplicados no próprio arquivo novo
    dupl_interno = df_novo.duplicated(subset=['ID _Ordem de Serviço'], keep='first').sum()
    if dupl_interno > 0:
        df_novo = df_novo.drop_duplicates(subset=['ID _Ordem de Serviço'], keep='first')
        log(f"Duplicados internos removidos: {dupl_interno}")

    # Verificar quais já existem na base
    ids_base = set(df_base['ID _Ordem de Serviço'].dropna().astype(str))
    ids_novo = df_novo['ID _Ordem de Serviço'].astype(str)
    ja_existem = ids_novo.isin(ids_base)

    novos = df_novo[~ja_existem]
    existentes = ja_existem.sum()

    log(f"OS já existentes na base (ignorados): {existentes}")
    log(f"OS novas para integrar: {len(novos)}")

    if len(novos) == 0:
        log("Nenhuma OS nova para integrar.")
        return df_base, 0

    # Garantir mesmas colunas
    for col in COLUNAS_OS:
        if col not in novos.columns:
            novos[col] = None

    novos = novos[COLUNAS_OS]
    df_integrado = pd.concat([df_base, novos], ignore_index=True)

    return df_integrado, len(novos)


def recalcular_relatorio(data_file):
    """Recalcula a aba RELATORIO replicando XLOOKUPs do Google Sheets.

    Para cada patrimônio na aba NOTAS, busca a última OS na aba OS.
    """
    log("Recalculando RELATORIO...")

    notas = pd.read_excel(data_file, sheet_name='NOTAS')
    os_df = pd.read_excel(data_file, sheet_name='OS')
    config = pd.read_excel(data_file, sheet_name='config')

    # Converter datas
    os_df['data_fechamento_OS'] = pd.to_datetime(os_df['data_fechamento_OS'], errors='coerce')

    # Limpar patrimônios
    os_df = os_df.dropna(subset=['id_patrimonio'])
    os_df['id_patrimonio'] = os_df['id_patrimonio'].astype(str).str.replace(r'\.0$', '', regex=True)

    notas_pat = notas.copy()
    if 'id_patrimonio' in notas_pat.columns:
        notas_pat['PATRIMONIO_STR'] = notas_pat['id_patrimonio'].astype(str).str.replace(r'\.0$', '', regex=True)
    else:
        print("ERRO: Coluna 'id_patrimonio' não encontrada na aba NOTAS.")
        return

    # Última OS por patrimônio
    os_sorted = os_df.sort_values('data_fechamento_OS')
    ultima_os = os_sorted.groupby('id_patrimonio').last().reset_index()

    # Mapeamento status_comodato → LOCAL_EQUIPAMENTO
    local_map = {
        'Emprestado': 'INSTALADO',
        'Sem Uso': 'EM ESTOQUE',
    }

    # Mapeamento almoxarifado → LOCAL_EQUIPAMENTO (para refinar)
    def inferir_local(row):
        status = str(row.get('STATUS COMODATO', ''))
        almox = str(row.get('ALMOXARIFADO', ''))

        if 'RMA' in almox.upper():
            return 'RMA'
        if 'Emprestado' in status:
            return 'INSTALADO'
        if 'Descontinuado' in almox or 'DESCONTINUADO' in almox.upper():
            return 'DESCONTINUADO'
        if any(x in almox.upper() for x in ['ALMOX', 'PRINCIPAL', 'DISTRIBUIC', 'CONFERIDO']):
            return 'EM ESTOQUE'
        if status == 'Sem Uso':
            return 'COM TÉCNICO'
        return 'COM TÉCNICO'

    # CICLO
    os_sorted_ciclo = os_df.sort_values(['id_patrimonio', 'data_fechamento_OS'])
    os_sorted_ciclo['CICLO'] = os_sorted_ciclo.groupby('id_patrimonio').cumcount() + 1
    ultimo_ciclo = os_sorted_ciclo.groupby('id_patrimonio')['CICLO'].last().reset_index()
    ultimo_ciclo.columns = ['id_patrimonio', 'ULTIMO_CICLO']

    # Construir RELATORIO
    rel = notas_pat.copy()
    rel = rel.rename(columns={
        'id_patrimonio': 'PATRIMONIO_ORIG',
        'Número NF': 'NF',
        'Data NF': 'DATA NF',
        'Descrição': 'DESCRICAO',
        'Nº Série': 'SERIE',
    })

    # Merge com última OS
    rel = rel.merge(
        ultima_os[['id_patrimonio', 'data_fechamento_OS', 'ASSUNTO PADRONIZADO',
                    'status_comodato', 'Almoxarifado']],
        left_on='PATRIMONIO_STR', right_on='id_patrimonio', how='left'
    )

    # Merge com CICLO
    rel = rel.merge(ultimo_ciclo, left_on='PATRIMONIO_STR', right_on='id_patrimonio', how='left')

    # Montar colunas finais
    rel_final = pd.DataFrame()
    rel_final['NF'] = rel['NF']
    rel_final['DATA NF'] = rel['DATA NF']
    rel_final['PRODUTO ID'] = rel.get('id_produto', None) if 'id_produto' in rel.columns else None
    rel_final['MAC'] = rel['MAC']
    rel_final['SERIE'] = rel['SERIE']
    rel_final['PATRIMONIO'] = rel['PATRIMONIO_STR']
    rel_final['DESCRICAO'] = rel['DESCRICAO']
    rel_final['ID CLIENTE'] = rel.get('ID_cliente', None) if 'ID_cliente' in rel.columns else None

    # ASSUNTO OS (padronizado da última OS, ou SEM OS)
    rel_final['ASSUNTO OS'] = rel['ASSUNTO PADRONIZADO'].fillna('SEM OS')
    # Aplicar padronização canônica
    rel_final['ASSUNTO OS'] = rel_final['ASSUNTO OS'].replace(PADRONIZACAO_ASSUNTO)

    rel_final['DATA ÚLTIMA OS'] = rel['data_fechamento_OS']
    rel_final['STATUS COMODATO'] = rel['status_comodato'].fillna('Sem Uso')
    rel_final['ALMOXARIFADO'] = rel['Almoxarifado'].fillna('')

    # STATUS_EQUIPAMENTO
    rel_final['STATUS_EQUIPAMENTO'] = rel['ULTIMO_CICLO'].apply(
        lambda x: 'REUTILIZADO' if pd.notna(x) and x > 1 else 'NOVO'
    )

    # LOCAL_EQUIPAMENTO
    rel_final['LOCAL_EQUIPAMENTO'] = rel_final.apply(inferir_local, axis=1)

    # Salvar no Excel (preservar outras abas)
    wb = openpyxl.load_workbook(data_file)

    # Remover aba RELATORIO existente
    if 'RELATORIO' in wb.sheetnames:
        del wb['RELATORIO']

    ws = wb.create_sheet('RELATORIO', 3)  # posição 3 (após NOTAS, OS)

    # Escrever headers
    headers = list(rel_final.columns)
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Escrever dados
    for row_idx, row_data in enumerate(rel_final.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            if pd.isna(value):
                ws.cell(row=row_idx, column=col_idx, value=None)
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(data_file)
    log(f"RELATORIO recalculado: {len(rel_final)} linhas")


def main():
    if len(sys.argv) < 2:
        print("Uso: python atualizar_mes.py <arquivo_os_do_mes>")
        print("Exemplo: python atualizar_mes.py OS_FEVEREIRO.xlsm")
        sys.exit(1)

    filepath = sys.argv[1]
    print(f"\n{'='*60}")
    print(f"  INGESTÃO MENSAL DE OS")
    print(f"  Arquivo: {filepath}")
    print(f"  Base: {DATA_FILE}")
    print(f"  Data: {datetime.now():%d/%m/%Y %H:%M:%S}")
    print(f"{'='*60}\n")

    # 1. Validar arquivo
    print("[1/6] Validando arquivo...")
    ext = validar_arquivo(filepath)

    # 2. Ler arquivo novo
    print("[2/6] Lendo arquivo de OS...")
    df_novo = ler_os_novo(filepath, ext)

    # 3. Carregar config e padronizar
    print("[3/6] Padronizando colunas...")
    config = pd.read_excel(DATA_FILE, sheet_name='config')
    df_novo = padronizar_colunas(df_novo, config)
    validar_colunas_obrigatorias(df_novo)

    # 4. Carregar base atual e integrar
    print("[4/6] Integrando com base existente...")
    df_base = pd.read_excel(DATA_FILE, sheet_name='OS')
    log(f"Base atual: {len(df_base)} linhas")

    df_integrado, qtd_novas = integrar_os(df_novo, df_base)

    if qtd_novas > 0:
        # 5. Salvar OS integrada
        print("[5/6] Salvando OS integrada...")
        wb = openpyxl.load_workbook(DATA_FILE)
        ws = wb['OS']

        # Limpar conteúdo existente (exceto header)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # Escrever dados integrados
        for row_idx, row_data in enumerate(df_integrado.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                if pd.isna(value):
                    ws.cell(row=row_idx, column=col_idx, value=None)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(DATA_FILE)
        log(f"OS salva: {len(df_integrado)} linhas (base anterior: {len(df_base)})")

        # 6. Recalcular RELATORIO
        print("[6/6] Recalculando RELATORIO...")
        recalcular_relatorio(DATA_FILE)
    else:
        print("[5/6] Nada para salvar.")
        print("[6/6] RELATORIO não precisa de recálculo.")

    # Relatório final
    print(f"\n{'='*60}")
    print(f"  INTEGRAÇÃO CONCLUÍDA")
    print(f"  OS novas integradas: {qtd_novas}")
    print(f"  Base final: {len(df_integrado)} linhas")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    main()
